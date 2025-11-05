import os
import openpyxl
import pandas as pd

# === USE PANDAS TO SCRAPE DATA ===
def scrape_data(FINAL_SAVE_DIR):
    # URL to scrape data from
    URL_DATA = "https://www.pro-football-reference.com/years/2025/"                                             

    # Read the data from the URL
    AFC = pd.read_html(URL_DATA, header=0)[0]
    NFC = pd.read_html(URL_DATA, header=0)[1]
    
    # Clean the data
    AFC = AFC.drop([0, 5, 10, 15])
    NFC = NFC.drop([0, 5, 10, 15])
    
    # combine the dataframes
    NFL = pd.concat([AFC, NFC], ignore_index=True)
    NFL = NFL.dropna()
    NFL = NFL.reset_index(drop=True)
    
    # if Team name has '*' or '+' remove it
    NFL['Tm'] = NFL['Tm'].str.replace('*', '', regex=False)
    NFL['Tm'] = NFL['Tm'].str.replace('+', '', regex=False)
    
    # Save the data to a CSV file to the final save directory
    os.makedirs(FINAL_SAVE_DIR, exist_ok=True)
    NFL.to_csv(os.path.join(FINAL_SAVE_DIR, "NFL_2025.csv"), index=False)

# === GET UPCOMING GAMES AND CREATE EXCEL FILE ===   
def get_upcoming_games(NFL_WEEK, PREDICTIONS_DIR):
    # URL to get weeklt matchups
    URL_GAMES = "https://www.pro-football-reference.com/years/2025/games.htm"
    
    # get the first table on the page, filter it
    GAMES = pd.read_html(URL_GAMES, header=0)[0]
    GAMES = GAMES[['Week', 'Winner/tie', 'Loser/tie']]
    GAMES = GAMES.rename(columns={'Winner/tie': 'Visitor', 'Loser/tie': 'Home'})
    week_num = NFL_WEEK.split()[1]
    GAMES = GAMES[GAMES['Week'] == week_num]
    GAMES = GAMES.reset_index(drop=True)
    
    # create an excel workbook with away team in column A and home team in column B, give headers 
    os.makedirs(PREDICTIONS_DIR, exist_ok=True)
    predictions = os.path.join(PREDICTIONS_DIR, f"{NFL_WEEK}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active

        # Write headers
    ws['A1'] = "Away Team"
    ws['B1'] = "Home Team"
    ws['C1'] = "Regression"
    ws['D1'] = "GBM"
    ws['E1'] = "Random Forest"
    ws['F1'] = "Stacking Ensemble"
    ws['G1'] = "Greg's Picks"

    # Write game data
    for idx, row in GAMES.iterrows():
        ws[f"A{idx+2}"] = row['Visitor']
        ws[f"B{idx+2}"] = row['Home']

    # save workbook
    wb.save(predictions)
    
# === MAIN FUNCTION ===
def main(NFL_WEEK, NFL_WEEK_MINUS_1):
    # === CONFIGURATION ===    
    FINAL_SAVE_DIR = "Data/" + NFL_WEEK_MINUS_1 
    PREDICTIONS_DIR = "Predictions/" + NFL_WEEK
    # Call the scrape_data function
    scrape_data(FINAL_SAVE_DIR)
    # Call the get_upcoming_games function
    get_upcoming_games(NFL_WEEK, PREDICTIONS_DIR)
