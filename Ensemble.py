import pickle
import openpyxl
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.svm import SVR
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import StackingRegressor
from sklearn.metrics import mean_squared_error, r2_score

# === NFL WEEK NUMBER ===
NFL_WEEK_TXT_FILE = "C:/Users/greg_yonan/Desktop/Tools/Programs/NFL/2025 Season/Week_Counter.txt"
with open(NFL_WEEK_TXT_FILE, 'r') as file:
    NFL_WEEK = file.read().strip()
    NFL_WEEK = int(NFL_WEEK.split()[1])
    NFL_WEEK = f"Week {NFL_WEEK}"
    NFL_WEEK_MINUS_1 = f"Week {int(NFL_WEEK.split()[1]) - 1}"

# File paths
DATA_FILENAME = "Data/" + NFL_WEEK_MINUS_1 + "/NFL_2025.csv"
PREDICTIONS_FILENAME = "Predictions/" + NFL_WEEK + '/' +  NFL_WEEK + ".xlsx"

# Below are column indicies
TEAM = 0 # Team Name
WINS = 1 # Wins
LOSSES = 2 # Losses
WIN_PERCENTAGE = 3 # Win Percentage
POINTS_FOR = 4 # Points For
POINTS_AGAINST = 5 # Points Against
POINTS_DIFFERENTIAL = 6 # Points Differential
MARGIN_OF_VICTORY = 7 # Margin of Victory
STRENGTH_OF_SCHEDULE = 8 # Strength of Schedule
SIMPLE_RATING_SYSTEM = 9 # Simple Rating System
OFFENSIVE_SRS = 10 # Offensive SRS
DEFENSIVE_SRS = 11 # Defensive SRS

# List of all teams
NFL_TEAMS = ['Arizona Cardinals', 'Atlanta Falcons', 'Baltimore Ravens', 'Buffalo Bills', 'Carolina Panthers', 'Chicago Bears', 'Cincinnati Bengals', 'Cleveland Browns', 'Dallas Cowboys', 'Denver Broncos', 'Detroit Lions', 'Green Bay Packers', 'Houston Texans', 'Indianapolis Colts', 'Jacksonville Jaguars', 'Kansas City Chiefs', 'Las Vegas Raiders', 'Los Angeles Chargers', 'Los Angeles Rams', 'Miami Dolphins', 'Minnesota Vikings', 'New England Patriots', 'New Orleans Saints', 'New York Giants', 'New York Jets', 'Philadelphia Eagles', 'Pittsburgh Steelers', 'San Francisco 49ers', 'Seattle Seahawks', 'Tampa Bay Buccaneers', 'Tennessee Titans', 'Washington Football Team']

params = {
    'stack_method': ['auto', 'predict_proba', 'decision_function', 'predict']
}

# Get team data     
def getTeamData(team_name):
    team_data = []
    with open(DATA_FILENAME, "r") as data_file:
        data = pd.read_csv(data_file)
        for index, row in data.iterrows():
            if row.iloc[TEAM] == team_name:
                team_data.append([row.iloc[WINS], row.iloc[LOSSES], row.iloc[WIN_PERCENTAGE], row.iloc[POINTS_FOR], row.iloc[POINTS_AGAINST], row.iloc[POINTS_DIFFERENTIAL], row.iloc[MARGIN_OF_VICTORY], row.iloc[STRENGTH_OF_SCHEDULE], row.iloc[SIMPLE_RATING_SYSTEM], row.iloc[OFFENSIVE_SRS], row.iloc[DEFENSIVE_SRS]])
            elif row.iloc[TEAM] == team_name:
                team_data.append([row.iloc[WINS], row.iloc[LOSSES], row.iloc[WIN_PERCENTAGE], row.iloc[POINTS_FOR], row.iloc[POINTS_AGAINST], row.iloc[POINTS_DIFFERENTIAL], row.iloc[MARGIN_OF_VICTORY], row.iloc[STRENGTH_OF_SCHEDULE], row.iloc[SIMPLE_RATING_SYSTEM], row.iloc[OFFENSIVE_SRS], row.iloc[DEFENSIVE_SRS]])
    team_data = pd.DataFrame(team_data, columns=['W', 'L', 'W-L%', 'PF', 'PA', 'PD', 'MoV', 'SoS', 'SRS', 'OSRS', 'DSRS'])
    return team_data 

# Train the model and save it to a file
def Train():
    NFL_team_data = []
    for team in NFL_TEAMS:
        team_data = getTeamData(team)
        if team_data is not None and not team_data.empty:  # Ensure team_data is not None or empty
            NFL_team_data.append(team_data)
    
    # Concatenate the list of DataFrames
    if NFL_team_data:  # Check if the list is not empty
        combined_data = pd.concat(NFL_team_data, ignore_index=True)
    else:
        raise ValueError("No valid team data found to train the model.")
    
    # Convert columns to numeric types
    combined_data['W'] = pd.to_numeric(combined_data['W'], errors='coerce')
    combined_data['L'] = pd.to_numeric(combined_data['L'], errors='coerce')
    combined_data['W-L%'] = pd.to_numeric(combined_data['W-L%'], errors='coerce')
    combined_data['PF'] = pd.to_numeric(combined_data['PF'], errors='coerce')
    combined_data['PA'] = pd.to_numeric(combined_data['PA'], errors='coerce')
    combined_data['PD'] = pd.to_numeric(combined_data['PD'], errors='coerce')
    combined_data['MoV'] = pd.to_numeric(combined_data['MoV'], errors='coerce')
    combined_data['SoS'] = pd.to_numeric(combined_data['SoS'], errors='coerce')
    combined_data['SRS'] = pd.to_numeric(combined_data['SRS'], errors='coerce')
    combined_data['OSRS'] = pd.to_numeric(combined_data['OSRS'], errors='coerce')
    combined_data['DSRS'] = pd.to_numeric(combined_data['DSRS'], errors='coerce')
    
   # Drop rows with any NaN values
    combined_data.dropna(subset=['W', 'L', 'W-L%', 'PF', 'PA', 'PD', 'MoV', 'SoS', 'SRS', 'OSRS', 'DSRS'], inplace=True)
    
    # Create features and labels
    x = combined_data[['W', 'L', 'PF', 'PA', 'PD', 'MoV', 'SoS', 'SRS', 'OSRS', 'DSRS']]
    y = combined_data['W-L%']
        
    # Split the data into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.3, random_state=42)
    
    # Define base models
    base_models = [
        ('dt', DecisionTreeRegressor()),
        ('mlp', MLPRegressor(max_iter=500)),
        ('svr', SVR())
    ]
    
    # Define the meta model
    meta_model = LinearRegression()
    
    # Define the stacking regressor
    model = StackingRegressor(estimators=base_models, final_estimator=meta_model)
    
    # Train the model
    model.fit(X_train, y_train)
    
    # Evaluate the model
    y_pred = model.predict(X_test)
    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)
    print(f'Mean Squared Error: {mse}')
    print(f'R^2 Score: {r2}') 
    
    # Save the model to disk
    MODEL_FILENAME = 'Models/' + NFL_WEEK + '/stacking_model.sav'
    pickle.dump(model, open(MODEL_FILENAME, 'wb'))
    print('Model saved to disk')

# Given a list of NFL games on an xlsx, predict the outcome of each game
def printOutcomes(model):
    # Read the Excel file    
    data = pd.read_excel(PREDICTIONS_FILENAME)
    data['Stacking Ensemble'] = data['Stacking Ensemble'].astype(str)
    
    for index, row in data.iterrows():
        team_a = row['Away Team']
        team_b = row['Home Team']
        team_a_data = getTeamData(team_a)
        team_b_data = getTeamData(team_b)
        team_a_features = pd.DataFrame([team_a_data[['W', 'L', 'PF', 'PA', 'PD', 'MoV', 'SoS', 'SRS', 'OSRS', 'DSRS']].mean().values], columns=['W', 'L', 'PF', 'PA', 'PD', 'MoV', 'SoS', 'SRS', 'OSRS', 'DSRS'])
        team_b_features = pd.DataFrame([team_b_data[['W', 'L', 'PF', 'PA', 'PD', 'MoV', 'SoS', 'SRS', 'OSRS', 'DSRS']].mean().values], columns=['W', 'L', 'PF', 'PA', 'PD', 'MoV', 'SoS', 'SRS', 'OSRS', 'DSRS'])
        team_a_win_prob = model.predict(team_a_features)[0]
        team_b_win_prob = model.predict(team_b_features)[0]
        
        if team_a_win_prob > team_b_win_prob:
            data.at[index, 'Stacking Ensemble'] = team_a
        else:
            data.at[index, 'Stacking Ensemble'] = team_b
    
    data.to_excel(PREDICTIONS_FILENAME, index=False)
    
def formatExcel():
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(PREDICTIONS_FILENAME)
    ws = wb.active
    
    # border around cells A1 to G1
    border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.border = border
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.font = openpyxl.styles.Font(bold=True)

    # cell A1 red
    red_fill = openpyxl.styles.PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
    ws['A1'].fill = red_fill
    # cell B1 blue
    blue_fill = openpyxl.styles.PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
    ws['B1'].fill = blue_fill
    # cell C1 yellow
    yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")
    ws['C1'].fill = yellow_fill
    # cell D1 purple
    purple_fill = openpyxl.styles.PatternFill(start_color="B1A0C7", end_color="B1A0C7", fill_type="solid")
    ws['D1'].fill = purple_fill
    # cell E1 green
    green_fill = openpyxl.styles.PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    ws['E1'].fill = green_fill
    # cell F1 orange
    orange_fill = openpyxl.styles.PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
    ws['F1'].fill = orange_fill
    # cell G1 gray
    gray_fill = openpyxl.styles.PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    ws['G1'].fill = gray_fill
    
    # auto adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width 
        
    # right boarder for cells A2 - G(number of rows), bottom border for cells A(number of rows) - G(number of rows) 
    for row in range(2, ws.max_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='thin'))     
            if row == ws.max_row:
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))       
            
    # Save the workbook
    wb.save(PREDICTIONS_FILENAME)

if __name__ == '__main__':
    # Train the model (uncomment if you need to train the model)
    Train()
    
    # Load the model and scaler
    model = pickle.load(open('Models/' + NFL_WEEK + '/stacking_model.sav', 'rb'))
    
    # Predict the outcomes
    printOutcomes(model)
    
    # Format the Excel file
    formatExcel()
    
    print("Ensemble Complete")